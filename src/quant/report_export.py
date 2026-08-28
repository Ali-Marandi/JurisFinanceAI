import numpy as np
import csv
import os
from datetime import datetime


class ReportExporter:
    """Export analysis results to PDF and Excel formats.

    Generates professional financial reports with tables,
    charts descriptions, and formatted output.
    Uses reportlab for PDF and openpyxl for Excel when available.
    Falls back to CSV when external libraries are not installed.
    """

    def export_to_csv(self, data, filepath, headers=None):
        """Export data to CSV file.

        Args:
            data: 2D numpy array or list of lists
            filepath: Output file path
            headers: Optional column headers
        """
        data = np.asarray(data)
        if data.ndim == 1:
            data = data.reshape(-1, 1)

        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            if headers:
                writer.writerow(headers)
            writer.writerows(data.tolist())

        return {'success': True, 'filepath': filepath, 'format': 'CSV', 'rows': data.shape[0]}

    def export_portfolio_report(self, filepath, analysis_results, format='csv'):
        """Export portfolio analysis results to a report file.

        Args:
            filepath: Output path (extension auto-adjusted)
            analysis_results: Dict from portfolio analysis
            format: 'csv' or 'txt'
        """
        if format == 'csv':
            return self._export_portfolio_csv(filepath, analysis_results)
        else:
            return self._export_portfolio_txt(filepath, analysis_results)

    def _export_portfolio_csv(self, filepath, analysis_results):
        rows = []
        headers = ['Metric', 'Value']

        if 'weights' in analysis_results:
            for name, w in zip(analysis_results.get('names', []), analysis_results['weights']):
                rows.append([f'Weight: {name}', f'{w:.4f}'])

        for key in ['expected_return', 'volatility', 'sharpe_ratio', 'sortino_ratio']:
            if key in analysis_results:
                label = key.replace('_', ' ').title()
                rows.append([label, f'{analysis_results[key]:.6f}'])

        if 'efficient_frontier' in analysis_results:
            rows.append([])
            rows.append(['--- Efficient Frontier ---', ''])
            ef = analysis_results['efficient_frontier']
            for ret, vol in zip(ef['returns'], ef['volatilities']):
                rows.append([f'Return', f'{ret:.6f}'])
                rows.append([f'Volatility', f'{vol:.6f}'])

        if not rows:
            rows.append(['No results', ''])

        filepath = filepath if filepath.endswith('.csv') else filepath + '.csv'
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

        return {'success': True, 'filepath': filepath, 'format': 'CSV', 'rows': len(rows)}

    def _export_portfolio_txt(self, filepath, analysis_results):
        lines = []
        lines.append('=' * 60)
        lines.append('JurisFinanceAI - Portfolio Analysis Report')
        lines.append(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        lines.append('=' * 60)
        lines.append('')

        if 'weights' in analysis_results:
            lines.append('Portfolio Weights:')
            for name, w in zip(analysis_results.get('names', []), analysis_results['weights']):
                lines.append(f'  {name}: {w:.2%}')
            lines.append('')

        for key in ['expected_return', 'volatility', 'sharpe_ratio', 'sortino_ratio']:
            if key in analysis_results:
                label = key.replace('_', ' ').title()
                lines.append(f'{label}: {analysis_results[key]:.6f}')
        lines.append('')

        if 'efficient_frontier' in analysis_results:
            lines.append('Efficient Frontier (10 points):')
            ef = analysis_results['efficient_frontier']
            for ret, vol in zip(ef['returns'][:10], ef['volatilities'][:10]):
                lines.append(f'  Return={ret:.4%}  Volatility={vol:.4%}  Sharpe={ret/(vol+1e-10):.3f}')

        filepath = filepath if filepath.endswith('.txt') else filepath + '.txt'
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        return {'success': True, 'filepath': filepath, 'format': 'TXT', 'rows': len(lines)}

    def export_risk_report(self, filepath, risk_results, format='csv'):
        """Export risk analysis results."""
        rows = []
        headers = ['Risk Metric', 'Value', 'Confidence']

        if 'var_results' in risk_results:
            for method, vals in risk_results['var_results'].items():
                if isinstance(vals, dict):
                    for cl, v in vals.items():
                        rows.append([f'VaR ({method})', f'{v:.6f}', cl])
                else:
                    rows.append([f'VaR ({method})', f'{vals:.6f}', '-'])

        for key in ['max_drawdown', 'altman_z', 'conditional_var']:
            if key in risk_results:
                label = key.replace('_', ' ').title()
                val = risk_results[key]
                if isinstance(val, (int, float)):
                    rows.append([label, f'{val:.6f}', '-'])

        if 'garch' in risk_results:
            g = risk_results['garch']
            rows.append(['GARCH Omega', f'{g.get(chr(111)+chr(109)+chr(101)+chr(103)+chr(97), 0):.6f}', '-'])
            rows.append(['GARCH Alpha', f'{g.get(chr(97)+chr(108)+chr(112)+chr(104)+chr(97), 0):.6f}', '-'])
            rows.append(['GARCH Beta', f'{g.get(chr(98)+chr(101)+chr(116)+chr(97), 0):.6f}', '-'])

        if not rows:
            rows.append(['No results', '', ''])

        filepath = filepath if filepath.endswith('.csv') else filepath + '.csv'
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

        return {'success': True, 'filepath': filepath, 'format': 'CSV', 'rows': len(rows)}

    def export_full_report(self, filepath, all_results, format='txt'):
        """Export comprehensive analysis report combining all results."""
        lines = []
        lines.append('=' * 70)
        lines.append('  JurisFinanceAI v3.0 - Comprehensive Analysis Report')
        lines.append(f'  Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        lines.append('=' * 70)

        section_count = 0
        for section_name, results in all_results.items():
            if not results or not isinstance(results, dict):
                continue
            section_count += 1
            lines.append('')
            lines.append('-' * 70)
            lines.append(f'  Section {section_count}: {section_name}')
            lines.append('-' * 70)

            for key, val in results.items():
                if isinstance(val, float):
                    lines.append(f'  {key}: {val:.6f}')
                elif isinstance(val, (int, bool, str)):
                    lines.append(f'  {key}: {val}')
                elif isinstance(val, np.ndarray):
                    lines.append(f'  {key}: array shape {val.shape}')
                elif isinstance(val, list) and len(val) > 0:
                    if isinstance(val[0], float):
                        lines.append(f'  {key}: [{val[0]:.4f}, ..., {val[-1]:.4f}] ({len(val)} items)')
                    else:
                        lines.append(f'  {key}: {len(val)} items')
                elif isinstance(val, dict):
                    for k2, v2 in val.items():
                        if isinstance(v2, float):
                            lines.append(f'  {key}.{k2}: {v2:.6f}')
                        else:
                            lines.append(f'  {key}.{k2}: {v2}')

        lines.append('')
        lines.append('=' * 70)
        lines.append(f'  Report End - {section_count} sections')
        lines.append('=' * 70)

        filepath = filepath if filepath.endswith(f'.{format}') else f'{filepath}.{format}'
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        return {'success': True, 'filepath': filepath, 'format': format.upper(), 'sections': section_count}

    def export_to_pdf(self, filepath, all_results):
        """Export report to PDF using reportlab if available."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.lib.colors import HexColor
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet

            doc = SimpleDocTemplate(filepath, pagesize=A4,
                                    rightMargin=20*mm, leftMargin=20*mm,
                                    topMargin=20*mm, bottomMargin=20*mm)
            styles = getSampleStyleSheet()
            elements = []

            # Title
            title_style = styles['Title']
            title_style.fontSize = 20
            title_style.textColor = HexColor('#1e40af')
            elements.append(Paragraph('JurisFinanceAI Analysis Report', title_style))
            elements.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
            elements.append(Spacer(1, 10*mm))

            for section_name, results in all_results.items():
                if not results or not isinstance(results, dict):
                    continue
                elements.append(Paragraph(f'<b>{section_name}</b>', styles['Heading2']))
                elements.append(Spacer(1, 3*mm))

                table_data = [['Metric', 'Value']]
                for key, val in results.items():
                    if isinstance(val, float):
                        table_data.append([key, f'{val:.6f}'])
                    elif isinstance(val, (int, str, bool)):
                        table_data.append([key, str(val)])

                if len(table_data) > 1:
                    t = Table(table_data, colWidths=[80*mm, 70*mm])
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1e40af')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#ffffff')),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#94a3b8')),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#f8fafc'), HexColor('#ffffff')]),
                    ]))
                    elements.append(t)
                elements.append(Spacer(1, 5*mm))

            doc.build(elements)
            return {'success': True, 'filepath': filepath, 'format': 'PDF'}

        except ImportError:
            # Fallback to text
            txt_path = filepath.replace('.pdf', '.txt')
            result = self.export_full_report(txt_path, all_results, format='txt')
            result['note'] = 'PDF library not available, exported as TXT'
            return result
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def export_to_excel(self, filepath, data_dict):
        """Export data to Excel using openpyxl if available."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()

            header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            cell_font = Font(name='Calibri', size=10)
            thin_border = Border(
                left=Side(style='thin', color='D1D5DB'),
                right=Side(style='thin', color='D1D5DB'),
                top=Side(style='thin', color='D1D5DB'),
                bottom=Side(style='thin', color='D1D5DB'),
            )

            for sheet_name, sheet_data in data_dict.items():
                if isinstance(sheet_data, dict):
                    rows_data = [[k, str(v) if not isinstance(v, float) else f'{v:.6f}']
                                for k, v in sheet_data.items()]
                elif isinstance(sheet_data, np.ndarray):
                    rows_data = sheet_data.tolist()
                elif isinstance(sheet_data, list):
                    rows_data = sheet_data
                else:
                    continue

                ws = wb.create_sheet(title=sheet_name[:31])

                for row_idx, row in enumerate(rows_data):
                    for col_idx, val in enumerate(row):
                        cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=val)
                        cell.font = header_font if row_idx == 0 else cell_font
                        cell.fill = header_fill if row_idx == 0 else PatternFill()
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')

                # Auto-width
                for col in ws.columns:
                    max_length = 0
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

            # Remove default sheet
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb['Sheet']

            wb.save(filepath)
            return {'success': True, 'filepath': filepath, 'format': 'Excel', 'sheets': len(data_dict)}

        except ImportError:
            csv_path = filepath.replace('.xlsx', '.csv')
            return self.export_to_csv(list(data_dict.values())[0] if data_dict else [], csv_path)
        except Exception as e:
            return {'success': False, 'error': str(e)}
